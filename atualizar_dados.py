#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Atualizador de dados — Fluxo de Caixa Centranet H&F (versão GitHub Actions)

Roda uma única vez por execução do workflow (agendado a cada 15 min).
Lê o arquivo no SharePoint via Microsoft Graph API (client credentials,
credenciais vindas de variáveis de ambiente / GitHub Secrets — nunca
gravadas em disco) e grava dados.json na raiz do repositório.

Este script NUNCA escreve no arquivo original do SharePoint (somente GET).
"""

import json, os, sys, re, unicodedata
from datetime import datetime, date
import requests
import openpyxl

GRAPH_BASE = "https://graph.microsoft.com/v1.0"
CACHE_PATH = "fluxo_caixa.xlsx"
OUT_PATH = "dados.json"

def env(name, required=True, default=None):
    v = os.environ.get(name, default)
    if required and not v:
        print(f"ERRO: variável de ambiente {name} não definida (configure em Settings > Secrets).",
              file=sys.stderr)
        sys.exit(1)
    return v

TENANT_ID = env("MS_TENANT_ID")
CLIENT_ID = env("MS_CLIENT_ID")
CLIENT_SECRET = env("MS_CLIENT_SECRET")
DRIVE_ID = env("MS_DRIVE_ID")
ITEM_ID = env("MS_ITEM_ID")
MIN_YEAR = int(env("MIN_YEAR", required=False, default="2026"))
MIN_MONTH = int(env("MIN_MONTH", required=False, default="1"))

# ─── Meses PT-BR ──────────────────────────────────────────────────────────────
PT_MONTHS = {
    "JANEIRO": 1, "FEVEREIRO": 2, "MARCO": 3, "ABRIL": 4, "MAIO": 5, "JUNHO": 6,
    "JULHO": 7, "AGOSTO": 8, "SETEMBRO": 9, "OUTUBRO": 10, "NOVEMBRO": 11, "DEZEMBRO": 12,
}

def _strip_accents(s):
    return "".join(c for c in unicodedata.normalize("NFD", s) if unicodedata.category(c) != "Mn")

def parse_sheet_name(name):
    if not name:
        return None
    norm = _strip_accents(name).strip().upper()
    m = re.match(r"^([A-Z]+)\s+(\d{4})$", norm)
    if not m:
        return None
    mes_txt, ano_txt = m.groups()
    mes_num = PT_MONTHS.get(mes_txt)
    if not mes_num:
        return None
    return int(ano_txt), mes_num

def to_float(v):
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    try:
        s = str(v).strip().replace("R$", "").replace(".", "").replace(",", ".")
        return float(s)
    except Exception:
        return None

def to_str(v):
    return str(v).strip() if v is not None else ""

def is_date_cell(v):
    return isinstance(v, (datetime, date)) and not isinstance(v, bool)

def as_date(v):
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    return None

def find_date_header_row(ws, ano, mes, max_scan_rows=8):
    best_row, best_map = None, {}
    for r in range(1, max_scan_rows + 1):
        col_map = {}
        for c in range(1, ws.max_column + 1):
            v = ws.cell(row=r, column=c).value
            if is_date_cell(v):
                d = as_date(v)
                if d and d.year == ano and d.month == mes:
                    col_map[c] = d
        if len(col_map) > len(best_map):
            best_row, best_map = r, col_map
    # Exige só 1 célula de data batendo com o mês/ano — o mês corrente pode ter
    # apenas 1-2 dias lançados até agora, e ainda assim deve ser reconhecido.
    if best_map and len(best_map) >= 1:
        return best_row, best_map
    return None, {}

def find_accounts_and_total(ws, date_header_row, max_scan_rows=80):
    accounts = []
    total_row = None
    start = date_header_row + 1
    for r in range(start, start + max_scan_rows):
        v = ws.cell(row=r, column=1).value
        txt = to_str(v)
        if not txt:
            continue
        norm = _strip_accents(txt).upper()
        if "SALDO" in norm and "TOTAL" in norm:
            total_row = r
            break
        accounts.append((r, txt))
    return accounts, total_row

def parse_sheet(ws, ano, mes):
    date_header_row, date_cols = find_date_header_row(ws, ano, mes)
    if date_header_row is None:
        return None, "linha de datas não identificada"
    accounts, total_row = find_accounts_and_total(ws, date_header_row)
    if not accounts:
        return None, "contas não identificadas"
    if total_row is None:
        return None, "linha SALDO TOTAL não encontrada"

    dias = []
    for col, d in sorted(date_cols.items(), key=lambda kv: kv[1]):
        contas_dia = {}
        for row, nome in accounts:
            val = to_float(ws.cell(row=row, column=col).value)
            contas_dia[nome] = val if val is not None else 0.0
        saldo_total = to_float(ws.cell(row=total_row, column=col).value)
        dias.append({
            "data": d.isoformat(),
            "saldo_total": saldo_total if saldo_total is not None else sum(contas_dia.values()),
            "contas": contas_dia,
        })
    return {
        "contas": [nome for _, nome in accounts],
        "dias": dias,
        "meta": {
            "date_header_row": date_header_row,
            "total_row": total_row,
            "linhas_contas": [r for r, _ in accounts],
        },
    }, None

def parse_workbook(path, min_year, min_month):
    wb = openpyxl.load_workbook(path, data_only=True)
    resultado_meses = []
    diagnostico_abas = []
    contas_set = []
    contas_seen = set()

    for sheet_name in wb.sheetnames:
        parsed_name = parse_sheet_name(sheet_name)
        diag = {"aba": sheet_name}
        if not parsed_name:
            diag["status"] = "ignorada (nome não reconhecido como mês/ano)"
            diagnostico_abas.append(diag)
            continue
        ano, mes = parsed_name
        if (ano, mes) < (min_year, min_month):
            diag["status"] = f"ignorada (anterior a {min_month:02d}/{min_year})"
            diagnostico_abas.append(diag)
            continue

        ws = wb[sheet_name]
        parsed, erro = parse_sheet(ws, ano, mes)
        if erro:
            diag["status"] = f"⚠️ A VERIFICAR — {erro}"
            diagnostico_abas.append(diag)
            continue

        diag["status"] = "ok"
        diag["dias_lidos"] = len(parsed["dias"])
        diagnostico_abas.append(diag)

        for nome in parsed["contas"]:
            if nome not in contas_seen:
                contas_seen.add(nome)
                contas_set.append(nome)

        resultado_meses.append({
            "mes": sheet_name, "ano": ano, "mes_num": mes, "dias": parsed["dias"],
        })

    wb.close()
    resultado_meses.sort(key=lambda m: (m["ano"], m["mes_num"]))
    return resultado_meses, contas_set, diagnostico_abas

def build_summary(meses):
    serie = []
    for m in meses:
        for d in m["dias"]:
            serie.append({"data": d["data"], "saldo_total": d["saldo_total"]})
    serie.sort(key=lambda x: x["data"])

    hoje = date.today().isoformat()
    passados = [d for d in serie if d["data"] <= hoje]
    ref = passados[-1] if passados else (serie[-1] if serie else None)

    saldo_por_conta_atual = {}
    if ref:
        for m in meses:
            for d in m["dias"]:
                if d["data"] == ref["data"]:
                    saldo_por_conta_atual = d["contas"]
                    break

    saldo_inicio_mes = None
    if ref:
        ref_ano, ref_mes = int(ref["data"][:4]), int(ref["data"][5:7])
        dias_mes = [d for d in serie if d["data"][:4] == str(ref_ano) and int(d["data"][5:7]) == ref_mes]
        if dias_mes:
            saldo_inicio_mes = dias_mes[0]["saldo_total"]

    variacao_mes = None
    if ref and saldo_inicio_mes is not None:
        variacao_mes = ref["saldo_total"] - saldo_inicio_mes

    return {
        "serie_diaria": serie,
        "saldo_atual": ref["saldo_total"] if ref else None,
        "data_saldo_atual": ref["data"] if ref else None,
        "saldo_por_conta_atual": saldo_por_conta_atual,
        "variacao_mes": variacao_mes,
    }

def get_token():
    url = f"https://login.microsoftonline.com/{TENANT_ID}/oauth2/v2.0/token"
    data = {
        "client_id": CLIENT_ID,
        "scope": "https://graph.microsoft.com/.default",
        "client_secret": CLIENT_SECRET,
        "grant_type": "client_credentials",
    }
    r = requests.post(url, data=data, timeout=20)
    if r.status_code != 200:
        print(f"ERRO ao obter token Graph ({r.status_code}): {r.text}", file=sys.stderr)
        sys.exit(1)
    return r.json()["access_token"]

def main():
    token = get_token()
    headers = {"Authorization": f"Bearer {token}"}

    meta_url = (f"{GRAPH_BASE}/drives/{DRIVE_ID}/items/{ITEM_ID}"
                f"?$select=id,name,lastModifiedDateTime,size,webUrl")
    r = requests.get(meta_url, headers=headers, timeout=20)
    if r.status_code != 200:
        print(f"ERRO ao consultar metadados ({r.status_code}): {r.text}", file=sys.stderr)
        sys.exit(1)
    meta = r.json()

    content_url = f"{GRAPH_BASE}/drives/{DRIVE_ID}/items/{ITEM_ID}/content"
    r = requests.get(content_url, headers=headers, timeout=60, allow_redirects=True)
    if r.status_code != 200:
        print(f"ERRO ao baixar conteúdo ({r.status_code}): {r.text[:300]}", file=sys.stderr)
        sys.exit(1)
    with open(CACHE_PATH, "wb") as f:
        f.write(r.content)

    meses, contas, diagnostico = parse_workbook(CACHE_PATH, MIN_YEAR, MIN_MONTH)
    resumo = build_summary(meses)

    dados = {
        "ok": True,
        "updated": datetime.utcnow().strftime("%d/%m/%Y às %H:%M:%S UTC"),
        "arquivo": {
            "nome": meta.get("name"),
            "last_modified": meta.get("lastModifiedDateTime"),
            "webUrl": meta.get("webUrl"),
            "size": meta.get("size"),
        },
        "contas": contas,
        "meses": meses,
        "resumo": resumo,
        "min_mes_ano": f"{MIN_MONTH:02d}/{MIN_YEAR}",
        "diagnostico_abas": diagnostico,
    }

    with open(OUT_PATH, "w", encoding="utf-8") as f:
        json.dump(dados, f, ensure_ascii=False, indent=2, default=str)

    os.remove(CACHE_PATH)  # não deixa a planilha original no runner além do necessário
    print(f"OK — {len(meses)} mês(es) lidos, {len(contas)} conta(s). Gravado em {OUT_PATH}")

if __name__ == "__main__":
    main()
